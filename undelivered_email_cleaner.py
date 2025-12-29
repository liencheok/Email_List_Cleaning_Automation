import pandas as pd
import re
import tkinter as tk
from tkinter import scrolledtext, messagebox, ttk, colorchooser
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ========== Email extraction function ==========
def extract_emails(text):
    emails = re.findall(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", text)
    return list(set([e.lower() for e in emails]))  # deduplicate, lowercase

# ========== Choose highlight color function ==========
def choose_color():
    color_code = colorchooser.askcolor(title="Choose Highlight Color")
    if color_code[1]:
        highlight_color_var.set(color_code[1].replace("#", "").upper())

# ========== Main processing function ==========
def process_emails():
    excel_file = file_textbox.get("1.0", tk.END).strip()
    undelivered_text = content_textbox.get("1.0", tk.END)
    highlight_color = highlight_color_var.get().strip() or "FFFF00"
    
    # Get selected list sheets
    list_selected_indices = list_sheet_listbox.curselection()
    list_sheet_names = [list_sheet_listbox.get(i) for i in list_selected_indices]
    
    # Get selected reject sheet
    reject_sheet_name = reject_sheet_var.get()
    
    if not excel_file:
        messagebox.showwarning("No File Path", "Please enter the Excel file path.")
        return
    if not list_sheet_names or not reject_sheet_name:
        messagebox.showwarning("Sheet Selection", "Please select at least one list sheet and a reject sheet.")
        return
    
    emails = extract_emails(undelivered_text)
    if not emails:
        messagebox.showwarning("No Emails Found", "No emails detected in pasted content.")
        return
    
    try:
        # Load reject sheet
        df_reject = pd.read_excel(excel_file, sheet_name=reject_sheet_name)
        
        moved_total = 0
        
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            for sheet in list_sheet_names:
                df_list = pd.read_excel(excel_file, sheet_name=sheet)
                email_col = df_list.columns[2]  # Email in third column
                df_list[email_col] = df_list[email_col].str.lower()
                
                # Find matching rows
                df_found = df_list[df_list[email_col].isin(emails)]
                
                if not df_found.empty:
                    moved_total += len(df_found)
                    
                    # Append to reject sheet
                    if df_reject.empty:
                        df_reject = pd.DataFrame(columns=df_list.columns)
                    else:
                        for col in df_list.columns:
                            if col not in df_reject.columns:
                                df_reject[col] = ""
                    
                    df_reject = pd.concat([df_reject, df_found], ignore_index=True)
                    
                    # Delete moved rows from list sheet
                    df_list = df_list[~df_list[email_col].isin(emails)]
                
                # Save updated list sheet
                df_list.to_excel(writer, sheet_name=sheet, index=False)
                
                # Set column width ~3.87 cm (25 units)
                worksheet = writer.sheets[sheet]
                for col in worksheet.columns:
                    col_letter = col[0].column_letter
                    worksheet.column_dimensions[col_letter].width = 25
            
            # Save updated reject sheet
            df_reject.to_excel(writer, sheet_name=reject_sheet_name, index=False)
            worksheet_reject = writer.sheets[reject_sheet_name]
            for col in worksheet_reject.columns:
                col_letter = col[0].column_letter
                worksheet_reject.column_dimensions[col_letter].width = 25
        
        # ========= Highlight Email Column in List Sheets ==========
        highlight_fill = PatternFill(start_color=highlight_color, end_color=highlight_color, fill_type="solid")
        wb = load_workbook(excel_file)
        for sheet in list_sheet_names:
            ws = wb[sheet]
            for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
                for cell in row:
                    cell.fill = highlight_fill
        wb.save(excel_file)
        
        # Completion message
        messagebox.showinfo(
            "Completed",
            f"✅ Moved total {moved_total} rows to \"{reject_sheet_name}\" from selected list sheets.\n"
            f"Email column highlighted with color #{highlight_color}."
        )
        
        # Ask if user wants to open Excel
        if messagebox.askyesno("Open Excel?", "Do you want to open the updated Excel file?"):
            os.startfile(excel_file)
        
        root.destroy()
    
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred:\n{e}")

# ========== Load sheet names ==========
def load_sheets():
    excel_file = file_textbox.get("1.0", tk.END).strip()
    if not excel_file:
        messagebox.showwarning("No File Path", "Please enter the Excel file path first.")
        return
    try:
        xls = pd.ExcelFile(excel_file)
        sheet_names = xls.sheet_names
        
        # Populate list sheet listbox
        list_sheet_listbox.delete(0, tk.END)
        for sheet in sheet_names:
            list_sheet_listbox.insert(tk.END, sheet)
        
        # Populate reject sheet combobox
        reject_sheet_dropdown['values'] = sheet_names
        
    except Exception as e:
        messagebox.showerror("Error", f"Failed to read sheets:\n{e}")

# ========== Tkinter GUI ==========
root = tk.Tk()
root.title("Undelivered Email Cleaner with Color Picker")

# Title bold
title_label = tk.Label(root, text="Undelivered Email Cleaner", font=("Arial", 12, "bold"))
title_label.pack(padx=10, pady=5)

tk.Label(root, text="Enter Excel file path:").pack(padx=10, pady=5)
file_textbox = scrolledtext.ScrolledText(root, width=100, height=2)
file_textbox.pack(padx=10, pady=5)

tk.Button(root, text="Load Sheets", command=load_sheets).pack(pady=5)

tk.Label(root, text="Select source (from) sheets:").pack(padx=10, pady=2)
list_sheet_listbox = tk.Listbox(root, selectmode='multiple', width=50, height=5)
list_sheet_listbox.pack(padx=10, pady=2)

reject_sheet_var = tk.StringVar()
tk.Label(root, text="Select reject (to) sheet:").pack(padx=10, pady=2)
reject_sheet_dropdown = ttk.Combobox(root, textvariable=reject_sheet_var, width=50)
reject_sheet_dropdown.pack(padx=10, pady=2)

# Highlight color picker
tk.Label(root, text="Highlight color (optional, default yellow):").pack(padx=10, pady=2)
highlight_color_var = tk.StringVar()
tk.Entry(root, textvariable=highlight_color_var, width=20).pack(padx=10, pady=2)
tk.Button(root, text="Choose Color", command=choose_color).pack(pady=2)

tk.Label(root, text="Paste undelivered email content below:").pack(padx=10, pady=5)
content_textbox = scrolledtext.ScrolledText(root, width=100, height=20)
content_textbox.pack(padx=10, pady=5)

tk.Button(root, text="Process Emails", command=process_emails).pack(pady=10)

root.mainloop()
